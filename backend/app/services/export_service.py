"""
Export Service: Generate CSV, Excel, and JSON exports of voter records.
Handles Bengali Unicode correctly in all export formats.
"""

import csv
import io
import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from sqlalchemy import and_, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.db.models import Document, VoterRecord
from app.schemas.schemas import ExportRequest

logger = logging.getLogger(__name__)

# All available voter record fields for export
ALL_EXPORT_FIELDS = [
    "id", "voter_id", "serial_number", "name", "name_english",
    "father_name", "mother_name", "spouse_name", "birth_date", "birth_year",
    "gender", "occupation", "address", "village", "post_office",
    "union_name", "ward", "upazila", "district", "division",
    "page_number", "pdf_file_name", "extraction_confidence",
]


async def fetch_export_records(
    db: AsyncSession,
    export_request: ExportRequest,
) -> List[Dict[str, Any]]:
    """
    Fetch voter records for export based on query/filter criteria.
    """
    stmt = (
        select(VoterRecord, Document.filename.label("pdf_filename"))
        .join(Document, VoterRecord.document_id == Document.id)
        .where(Document.status == "completed")
    )

    # Apply text filter if provided
    if export_request.query:
        pattern = f"%{export_request.query}%"
        stmt = stmt.where(
            or_(
                VoterRecord.name.ilike(pattern),
                VoterRecord.father_name.ilike(pattern),
                VoterRecord.voter_id.ilike(pattern),
                VoterRecord.district.ilike(pattern),
            )
        )

    # Apply additional filters
    if export_request.filters:
        filters = export_request.filters
        filter_map = {
            "district": VoterRecord.district,
            "upazila": VoterRecord.upazila,
            "union_name": VoterRecord.union_name,
            "ward": VoterRecord.ward,
            "gender": VoterRecord.gender,
        }
        for key, column in filter_map.items():
            if filters.get(key):
                stmt = stmt.where(column.ilike(f"%{filters[key]}%"))

    stmt = stmt.limit(export_request.max_records)

    result = await db.execute(stmt)
    rows = result.all()

    records = []
    for row in rows:
        record = row[0]
        filename = row[1]
        item = {
            "id": record.id,
            "voter_id": record.voter_id or "",
            "serial_number": record.serial_number or "",
            "name": record.name or "",
            "name_english": record.name_english or "",
            "father_name": record.father_name or "",
            "mother_name": record.mother_name or "",
            "spouse_name": record.spouse_name or "",
            "birth_date": record.birth_date or "",
            "birth_year": record.birth_year or "",
            "gender": record.gender or "",
            "occupation": record.occupation or "",
            "address": record.address or "",
            "village": record.village or "",
            "post_office": record.post_office or "",
            "union_name": record.union_name or "",
            "ward": record.ward or "",
            "upazila": record.upazila or "",
            "district": record.district or "",
            "division": record.division or "",
            "page_number": record.page_number or "",
            "pdf_file_name": filename or "",
            "extraction_confidence": record.extraction_confidence or "",
        }
        # Filter to requested fields only
        if export_request.fields:
            item = {k: v for k, v in item.items() if k in export_request.fields}
        records.append(item)

    return records


def export_to_csv(records: List[Dict[str, Any]], fields: Optional[List[str]] = None) -> bytes:
    """
    Export records to CSV bytes with UTF-8 BOM for Bengali Unicode support in Excel.
    
    Returns:
        UTF-8 BOM encoded CSV bytes.
    """
    if not records:
        return b"\xef\xbb\xbf"  # Empty CSV with BOM

    fieldnames = fields or (list(records[0].keys()) if records else ALL_EXPORT_FIELDS)

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=fieldnames,
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerows(records)

    # Add UTF-8 BOM for proper Bengali display in Excel
    return "\ufeff".encode("utf-8") + output.getvalue().encode("utf-8")


def export_to_excel(records: List[Dict[str, Any]], fields: Optional[List[str]] = None) -> bytes:
    """
    Export records to Excel (.xlsx) with proper Unicode support and styling.
    
    Returns:
        Excel file as bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        raise RuntimeError("openpyxl required for Excel export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Voter Records"

    fieldnames = fields or (list(records[0].keys()) if records else ALL_EXPORT_FIELDS)

    # Header row with styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field.upper().replace("_", " "))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            value = record.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.alignment = Alignment(horizontal="right" if _is_bengali_field(field) else "left")

    # Auto-adjust column widths
    for col_idx, field in enumerate(fieldnames, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(field)),
            max((len(str(r.get(field, "") or "")) for r in records), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze top row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_to_json(records: List[Dict[str, Any]]) -> bytes:
    """Export records to JSON with proper Unicode encoding."""
    return json.dumps(
        {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "total_records": len(records),
            "records": records,
        },
        ensure_ascii=False,
        indent=2,
    ).encode("utf-8")


def _is_bengali_field(field_name: str) -> bool:
    """Heuristic to determine if a field likely contains Bengali text."""
    bengali_fields = {
        "name", "father_name", "mother_name", "spouse_name",
        "address", "village", "post_office", "union_name",
        "upazila", "district", "division", "occupation",
    }
    return field_name in bengali_fields


async def generate_export(
    db: AsyncSession,
    export_request: ExportRequest,
) -> Dict[str, Any]:
    """
    Main export function: fetch data and serialize to requested format.
    
    Returns:
        {"content": bytes, "filename": str, "media_type": str}
    """
    logger.info(f"Generating {export_request.format} export (max {export_request.max_records} records)")

    records = await fetch_export_records(db, export_request)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if export_request.format == "csv":
        content = export_to_csv(records, export_request.fields)
        filename = f"voter_records_{timestamp}.csv"
        media_type = "text/csv; charset=utf-8"

    elif export_request.format == "excel":
        content = export_to_excel(records, export_request.fields)
        filename = f"voter_records_{timestamp}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    elif export_request.format == "json":
        content = export_to_json(records)
        filename = f"voter_records_{timestamp}.json"
        media_type = "application/json; charset=utf-8"

    else:
        raise ValueError(f"Unsupported export format: {export_request.format}")

    # Optionally save to disk
    export_path = Path(settings.EXPORT_DIR) / filename
    export_path.write_bytes(content)
    logger.info(f"Export saved to {export_path}: {len(records)} records")

    return {
        "content": content,
        "filename": filename,
        "media_type": media_type,
        "record_count": len(records),
    }
